"""MD -> XLSX. Extracts every pipe-table as its own sheet."""
from __future__ import annotations

from pathlib import Path

from markdown_it import MarkdownIt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import themes


def _extract_tables(md_text: str) -> list[tuple[list[str], list[list[str]]]]:
    """Return list of (headers, rows) from every table in the MD."""
    md = MarkdownIt("commonmark").enable("table")
    tokens = md.parse(md_text)
    tables = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if tok.type == "table_open":
            headers: list[str] = []
            rows: list[list[str]] = []
            current_row: list[str] = []
            in_head = False
            j = i + 1
            while j < len(tokens) and tokens[j].type != "table_close":
                t = tokens[j]
                if t.type == "thead_open":
                    in_head = True
                elif t.type == "thead_close":
                    in_head = False
                elif t.type == "tr_open":
                    current_row = []
                elif t.type == "tr_close":
                    if in_head:
                        headers = current_row
                    else:
                        rows.append(current_row)
                elif t.type == "inline":
                    current_row.append(t.content)
                j += 1
            tables.append((headers, rows))
            i = j
        i += 1
    return tables


def convert(src: Path, out: Path, assets_dir: Path, style: dict) -> None:  # noqa: ARG001
    text = src.read_text(encoding="utf-8")
    tables = _extract_tables(text)
    if not tables:
        raise RuntimeError("该文件中没有找到表格")

    font_id = style.get("font", themes.DEFAULTS["font"])
    size_id = style.get("size", themes.DEFAULTS["size"])
    color_id = style.get("color", themes.DEFAULTS["color"])
    primary = style.get("primary", "zh")

    s = themes.get_size(size_id)
    c = themes.get_color(color_id)

    body_font_name = themes.primary_name(font_id, "body", primary)
    body_pt = float(s["body_pt"])  # type: ignore[arg-type]

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name=body_font_name, size=body_pt, bold=True, color="FFFFFF")
    body_font = Font(name=body_font_name, size=body_pt - 0.5)
    header_fill = PatternFill("solid", fgColor=c["heading"])
    alt_fill = PatternFill("solid", fgColor=c["code_bg"])
    thin = Side(border_style="thin", color=c["rule"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, (headers, rows) in enumerate(tables, 1):
        ws = wb.create_sheet(title=f"Table {idx}"[:31])
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for r_i, row in enumerate(rows, 2):
            for c_i, val in enumerate(row, 1):
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.font = body_font
                cell.alignment = center
                cell.border = border
                if r_i % 2 == 0:
                    cell.fill = alt_fill

        for col_i, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(r[col_i - 1])) if col_i - 1 < len(r) else 0 for r in rows])
            ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = min(max(max_len + 4, 10), 60)
        ws.row_dimensions[1].height = 22

    wb.save(out)
